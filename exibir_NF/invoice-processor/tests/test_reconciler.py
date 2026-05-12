"""
tests/test_reconciler.py — Unit Tests for reconciler.py
---------------------------------------------------------
Tests all 5 public functions in reconciler.py:
    - reconcile()
    - read_pdf_total()
    - read_spreadsheet_total()
    - compare_totals()
    - report_result()

Run:
    python -m pytest tests/test_reconciler.py -v
"""

import sys
from dataclasses import asdict
from pathlib import Path

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
sys.path.insert(0, str(Path(__file__).parent.parent))

from reconciler import (
    STATUS_DIVERGENCE,
    STATUS_ERROR,
    STATUS_MATCH,
    ReconciliationResult,
    compare_totals,
    read_pdf_total,
    read_spreadsheet_total,
    reconcile,
    report_result,
)


# ─────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────

def _make_xlsx_with_total(tmp_path: Path, total: float, sheet_name: str = "MAIO 2025") -> Path:
    """Creates a minimal .xlsx with a numeric value in column E row 5."""
    path = tmp_path / "card_test.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Write a value in column E (VALOR) at row 5
    ws.cell(row=5, column=5, value=total)
    # Write a SUB TOTAL label so read_spreadsheet_total can find it
    ws.cell(row=6, column=4, value="SUB TOTAL")
    ws.cell(row=6, column=5, value=total)
    wb.save(str(path))
    return path


def _make_result(
    status: str = STATUS_MATCH,
    pdf_total: float = 1337.61,
    spreadsheet_total: float = 1337.61,
    difference: float = 0.0,
) -> ReconciliationResult:
    return ReconciliationResult(
        person_name="John Doe",
        sheet_name="MAIO 2025",
        spreadsheet_path="spreadsheets/card_john.xlsx",
        pdf_total=pdf_total,
        spreadsheet_total=spreadsheet_total,
        difference=difference,
        status=status,
        message="Test message",
        timestamp="2025-05-01T14:00:00",
    )


# ─────────────────────────────────────────────
# read_pdf_total
# ─────────────────────────────────────────────

class TestReadPdfTotal:

    def test_valid_positive_total(self):
        assert read_pdf_total(1337.61) == 1337.61

    def test_zero_total_allowed(self):
        result = read_pdf_total(0.0)
        assert result == 0.0

    def test_rounds_to_two_decimals(self):
        result = read_pdf_total(1337.6099999)
        assert result == 1337.61

    def test_negative_total_raises(self):
        with pytest.raises(ValueError, match="negative"):
            read_pdf_total(-100.00)

    def test_nan_raises(self):
        with pytest.raises(ValueError, match="NaN"):
            read_pdf_total(float("nan"))

    def test_non_numeric_raises(self):
        with pytest.raises(ValueError):
            read_pdf_total("1337.61")

    def test_exceeds_limit_raises(self):
        with pytest.raises(ValueError, match="safety limit"):
            read_pdf_total(2_000_000.00)

    def test_integer_input_accepted(self):
        result = read_pdf_total(1000)
        assert result == 1000.0


# ─────────────────────────────────────────────
# compare_totals
# ─────────────────────────────────────────────

class TestCompareTotals:

    def test_exact_match(self):
        status, message = compare_totals("John", 1337.61, 1337.61, 0.0)
        assert status == STATUS_MATCH
        assert "MATCH" in message

    def test_within_tolerance_is_match(self):
        status, _ = compare_totals("John", 1337.61, 1337.60, 0.01)
        assert status == STATUS_MATCH

    def test_divergence_detected(self):
        status, message = compare_totals("John", 1337.61, 1287.61, 50.0)
        assert status == STATUS_DIVERGENCE
        assert "DIVERGENCE" in message

    def test_message_contains_person_name(self):
        _, message = compare_totals("Aline Maris", 500.0, 500.0, 0.0)
        assert "Aline Maris" in message

    def test_message_contains_values(self):
        _, message = compare_totals("John", 500.0, 450.0, 50.0)
        assert "500" in message
        assert "450" in message

    def test_invalid_pdf_total_returns_error(self):
        status, _ = compare_totals("John", "bad", 500.0, 0.0)
        assert status == STATUS_ERROR

    def test_invalid_spreadsheet_total_returns_error(self):
        status, _ = compare_totals("John", 500.0, "bad", 0.0)
        assert status == STATUS_ERROR

    def test_zero_difference_is_match(self):
        status, _ = compare_totals("John", 0.0, 0.0, 0.0)
        assert status == STATUS_MATCH


# ─────────────────────────────────────────────
# read_spreadsheet_total
# ─────────────────────────────────────────────

class TestReadSpreadsheetTotal:

    def test_reads_subtotal_from_sheet(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 1337.61)
        total = read_spreadsheet_total(str(path), "MAIO 2025")
        assert total == 1337.61

    def test_file_not_found_raises(self):
        with pytest.raises(FileNotFoundError):
            read_spreadsheet_total("/nonexistent/file.xlsx", "MAIO 2025")

    def test_wrong_extension_raises(self, tmp_path):
        txt = tmp_path / "file.txt"
        txt.write_text("content")
        with pytest.raises(ValueError):
            read_spreadsheet_total(str(txt), "MAIO 2025")

    def test_sheet_not_found_raises(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 100.0)
        with pytest.raises(ValueError, match="not found"):
            read_spreadsheet_total(str(path), "JUNHO 2025")

    def test_returns_float(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 500.50)
        total = read_spreadsheet_total(str(path), "MAIO 2025")
        assert isinstance(total, float)


# ─────────────────────────────────────────────
# report_result
# ─────────────────────────────────────────────

class TestReportResult:

    def test_match_result_does_not_raise(self, capsys):
        result = _make_result(STATUS_MATCH, 1337.61, 1337.61, 0.0)
        report_result(result)
        captured = capsys.readouterr()
        assert "MATCH" in captured.out

    def test_divergence_result_does_not_raise(self, capsys):
        result = _make_result(STATUS_DIVERGENCE, 1337.61, 1287.61, 50.0)
        report_result(result)
        captured = capsys.readouterr()
        assert "DIVERGENCE" in captured.out

    def test_output_contains_person_name(self, capsys):
        result = _make_result()
        report_result(result)
        captured = capsys.readouterr()
        assert "John Doe" in captured.out

    def test_output_contains_pdf_total(self, capsys):
        result = _make_result(pdf_total=1337.61)
        report_result(result)
        captured = capsys.readouterr()
        assert "1,337.61" in captured.out or "1337" in captured.out

    def test_invalid_input_does_not_raise(self):
        report_result("not a result")  # should log error, not raise

    def test_output_contains_timestamp(self, capsys):
        result = _make_result()
        report_result(result)
        captured = capsys.readouterr()
        assert "2025-05-01" in captured.out


# ─────────────────────────────────────────────
# reconcile (integration)
# ─────────────────────────────────────────────

class TestReconcile:

    def test_match_result(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 1337.61)
        result = reconcile("John Doe", "MAIO 2025", str(path), 1337.61)
        assert result.status == STATUS_MATCH
        assert result.difference <= 0.01

    def test_divergence_result(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 1000.00)
        result = reconcile("John Doe", "MAIO 2025", str(path), 1337.61)
        assert result.status == STATUS_DIVERGENCE
        assert result.difference > 0.01

    def test_result_contains_person_name(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 500.0)
        result = reconcile("Aline Maris", "MAIO 2025", str(path), 500.0)
        assert result.person_name == "Aline Maris"

    def test_result_contains_correct_totals(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 850.0)
        result = reconcile("John", "MAIO 2025", str(path), 850.0)
        assert result.pdf_total == 850.0
        assert result.spreadsheet_total == 850.0

    def test_file_not_found_raises(self):
        with pytest.raises(FileNotFoundError):
            reconcile("John", "MAIO 2025", "/nonexistent/file.xlsx", 100.0)

    def test_empty_person_name_raises(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 100.0)
        with pytest.raises(ValueError):
            reconcile("", "MAIO 2025", str(path), 100.0)

    def test_negative_pdf_total_raises(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 100.0)
        with pytest.raises(ValueError):
            reconcile("John", "MAIO 2025", str(path), -100.0)

    def test_result_is_dataclass(self, tmp_path):
        path = _make_xlsx_with_total(tmp_path, 200.0)
        result = reconcile("John", "MAIO 2025", str(path), 200.0)
        assert isinstance(result, ReconciliationResult)
        assert "person_name" in asdict(result)