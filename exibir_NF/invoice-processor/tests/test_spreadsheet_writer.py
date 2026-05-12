"""
tests/test_spreadsheet_writer.py — Unit Tests for spreadsheet_writer.py
-------------------------------------------------------------------------
Tests all 5 public functions in spreadsheet_writer.py:
    - write_transactions()
    - open_spreadsheet()
    - locate_sheet()
    - insert_rows()
    - format_cells()

Run:
    python -m pytest tests/test_spreadsheet_writer.py -v
"""

import shutil
import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
sys.path.insert(0, str(Path(__file__).parent.parent))

from pdf_reader import Transaction, BANK_ITAU
from spreadsheet_writer import (
    format_cells,
    insert_rows,
    locate_sheet,
    open_spreadsheet,
    write_transactions,
    PAYMENT_METHOD,
    DATA_START_ROW,
    COL_DATE,
    COL_DESCRIPTION,
    COL_AMOUNT,
    COL_PAYMENT_METHOD,
)


# ─────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────

def _make_xlsx(tmp_path: Path, sheet_name: str = "MAIO 2025") -> Path:
    """Creates a minimal .xlsx file with the correct sheet name."""
    path = tmp_path / "test_card.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    # Add column headers at row 4
    headers = ["DATA", "NF/CUPOM", "FORNECEDOR", "DESCRIÇÃO",
               "VALOR", "FORMA DE PGTO", "PARCELAS", "PROJETO",
               "REEMBOLSÁVEL", "GESTOR", "OBSERVAÇÃO"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=4, column=col, value=header)
    wb.save(str(path))
    return path


def _make_transactions(count: int = 3) -> list[Transaction]:
    """Creates a list of sample Transaction objects."""
    return [
        Transaction(
            date=f"0{i}/10",
            description=f"STORE_{i}",
            amount=float(10 * i),
            bank=BANK_ITAU,
        )
        for i in range(1, count + 1)
    ]


# ─────────────────────────────────────────────
# open_spreadsheet
# ─────────────────────────────────────────────

class TestOpenSpreadsheet:

    def test_opens_valid_xlsx(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        assert wb is not None

    def test_file_not_found_raises(self):
        with pytest.raises(FileNotFoundError):
            open_spreadsheet("/nonexistent/file.xlsx")

    def test_wrong_extension_raises(self, tmp_path):
        txt = tmp_path / "file.txt"
        txt.write_text("content")
        with pytest.raises(ValueError, match="Invalid file type"):
            open_spreadsheet(str(txt))

    def test_empty_file_raises(self, tmp_path):
        empty = tmp_path / "empty.xlsx"
        empty.write_bytes(b"")
        with pytest.raises(ValueError):
            open_spreadsheet(str(empty))

    def test_oversized_file_raises(self, tmp_path):
        big = tmp_path / "big.xlsx"
        big.write_bytes(b"PK" + b"x" * (21 * 1024 * 1024))
        with pytest.raises(ValueError, match="exceeds"):
            open_spreadsheet(str(big))


# ─────────────────────────────────────────────
# locate_sheet
# ─────────────────────────────────────────────

class TestLocateSheet:

    def test_finds_exact_match(self, tmp_path):
        path = _make_xlsx(tmp_path, sheet_name="MAIO 2025")
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        assert ws.title == "MAIO 2025"

    def test_finds_case_insensitive_match(self, tmp_path):
        path = _make_xlsx(tmp_path, sheet_name="MAIO 2025")
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "maio 2025")
        assert ws is not None

    def test_raises_when_not_found(self, tmp_path):
        path = _make_xlsx(tmp_path, sheet_name="MAIO 2025")
        wb = open_spreadsheet(str(path))
        with pytest.raises(ValueError, match="not found"):
            locate_sheet(wb, "JUNHO 2025")

    def test_raises_on_empty_sheet_name(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        with pytest.raises(ValueError):
            locate_sheet(wb, "")

    def test_raises_on_none_sheet_name(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        with pytest.raises(ValueError):
            locate_sheet(wb, None)


# ─────────────────────────────────────────────
# insert_rows
# ─────────────────────────────────────────────

class TestInsertRows:

    def test_inserts_correct_number_of_rows(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        transactions = _make_transactions(5)
        written = insert_rows(ws, transactions, DATA_START_ROW)
        assert written == 5

    def test_returns_zero_for_empty_list(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        written = insert_rows(ws, [], DATA_START_ROW)
        assert written == 0

    def test_payment_method_prefilled(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        transactions = _make_transactions(1)
        insert_rows(ws, transactions, DATA_START_ROW)
        cell_value = ws.cell(row=DATA_START_ROW, column=COL_PAYMENT_METHOD).value
        assert cell_value == PAYMENT_METHOD

    def test_date_written_correctly(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        transactions = _make_transactions(1)
        insert_rows(ws, transactions, DATA_START_ROW)
        assert ws.cell(row=DATA_START_ROW, column=COL_DATE).value == "01/10"

    def test_amount_written_correctly(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        transactions = _make_transactions(1)
        insert_rows(ws, transactions, DATA_START_ROW)
        assert ws.cell(row=DATA_START_ROW, column=COL_AMOUNT).value == 10.0

    def test_start_row_before_data_area_raises(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        with pytest.raises(ValueError, match="start_row"):
            insert_rows(ws, _make_transactions(1), start_row=1)

    def test_description_written_correctly(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        transactions = _make_transactions(1)
        insert_rows(ws, transactions, DATA_START_ROW)
        desc = ws.cell(row=DATA_START_ROW, column=COL_DESCRIPTION).value
        assert desc == "STORE_1"


# ─────────────────────────────────────────────
# format_cells
# ─────────────────────────────────────────────

class TestFormatCells:

    def test_format_cells_does_not_raise(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        transactions = _make_transactions(3)
        insert_rows(ws, transactions, DATA_START_ROW)
        format_cells(ws, DATA_START_ROW, 3)  # should not raise

    def test_format_cells_zero_rows_does_not_raise(self, tmp_path):
        path = _make_xlsx(tmp_path)
        wb = open_spreadsheet(str(path))
        ws = locate_sheet(wb, "MAIO 2025")
        format_cells(ws, DATA_START_ROW, 0)  # should not raise


# ─────────────────────────────────────────────
# write_transactions (integration)
# ─────────────────────────────────────────────

class TestWriteTransactions:

    def test_full_write_pipeline(self, tmp_path):
        path = _make_xlsx(tmp_path)
        transactions = _make_transactions(3)
        written = write_transactions(str(path), "MAIO 2025", transactions)
        assert written == 3

    def test_file_saved_after_write(self, tmp_path):
        path = _make_xlsx(tmp_path)
        transactions = _make_transactions(2)
        write_transactions(str(path), "MAIO 2025", transactions)
        wb = load_workbook(str(path))
        ws = wb["MAIO 2025"]
        assert ws.cell(row=DATA_START_ROW, column=COL_DATE).value is not None

    def test_backup_created(self, tmp_path):
        path = _make_xlsx(tmp_path)
        write_transactions(str(path), "MAIO 2025", _make_transactions(1))
        backups = list(tmp_path.glob("*.backup.xlsx"))
        assert len(backups) == 1

    def test_file_not_found_raises(self):
        with pytest.raises(FileNotFoundError):
            write_transactions("/nonexistent/file.xlsx", "MAIO 2025", [])

    def test_empty_sheet_name_raises(self, tmp_path):
        path = _make_xlsx(tmp_path)
        with pytest.raises(ValueError):
            write_transactions(str(path), "", [])

    def test_too_many_transactions_raises(self, tmp_path):
        path = _make_xlsx(tmp_path)
        too_many = _make_transactions(1) * 501
        with pytest.raises(ValueError, match="Too many"):
            write_transactions(str(path), "MAIO 2025", too_many)