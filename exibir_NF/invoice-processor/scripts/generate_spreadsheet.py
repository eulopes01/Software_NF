"""
generate_spreadsheet.py — Phase 1: Spreadsheet Generator
----------------------------------------------------------
Generates one blank .xlsx file per person with the correct
structure for corporate credit card invoice control.

Automatically filled by the processing pipeline:
    DATA | FORNECEDOR | DESCRIÇÃO | VALOR | FORMA DE PGTO

Left blank for manual entry:
    NF/CUPOM | PARCELAS | PROJETO | REEMBOLSÁVEL | GESTOR | OBSERVAÇÃO

Run:
    python scripts/generate_spreadsheet.py

Output:
    One .xlsx file per person saved to the spreadsheets/ directory.

Dependencies:
    pip install openpyxl
"""

import logging
import sys
from pathlib import Path
from typing import NamedTuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Logging setup
# ─────────────────────────────────────────────

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Configuration — edit before running
# ─────────────────────────────────────────────

PEOPLE: list[str] = [
    "Aline Maris",
    "Carlos Belruss",
    "Daniela Sa",
    # Add or remove names as needed
]

REFERENCE_MONTH  = "MAIO 2025"    # Sheet tab name and header label
DUE_DATE         = "15/06/2025"   # Invoice due date shown in the header
OUTPUT_DIR       = Path("spreadsheets")
TRANSACTION_ROWS = 50             # Blank rows for national transactions
INTL_ROWS        = 5              # Blank rows for international transactions
ANNUITY_AMOUNT   = 18.75          # Default monthly annuity installment (R$)

# ─────────────────────────────────────────────
# Column definitions
# ─────────────────────────────────────────────

class ColumnDefinition(NamedTuple):
    header: str
    width:  float

COLUMNS: list[ColumnDefinition] = [
    ColumnDefinition("DATA",          12),
    ColumnDefinition("NF/CUPOM",      12),
    ColumnDefinition("FORNECEDOR",    28),
    ColumnDefinition("DESCRIÇÃO",     45),
    ColumnDefinition("VALOR",         14),
    ColumnDefinition("FORMA DE PGTO", 14),
    ColumnDefinition("PARCELAS",      14),
    ColumnDefinition("PROJETO",       10),
    ColumnDefinition("REEMBOLSÁVEL",  13),
    ColumnDefinition("GESTOR",        14),
    ColumnDefinition("OBSERVAÇÃO",    30),
]

TOTAL_COLUMNS = len(COLUMNS)

# ─────────────────────────────────────────────
# Style constants
# ─────────────────────────────────────────────

COLOR_DARK_BLUE   = "1F4E79"
COLOR_MED_BLUE    = "2E75B6"
COLOR_LIGHT_BLUE  = "D6E4F0"
COLOR_TOTAL_BLUE  = "BDD7EE"
COLOR_WHITE       = "FFFFFF"

FONT_TITLE        = Font(name="Arial", bold=True, color=COLOR_WHITE,    size=11)
FONT_HEADER_COL   = Font(name="Arial", bold=True, color=COLOR_WHITE,    size=10)
FONT_BOLD         = Font(name="Arial", bold=True,                       size=10)
FONT_NORMAL       = Font(name="Arial",                                  size=10)
FONT_TOTAL        = Font(name="Arial", bold=True, color=COLOR_WHITE,    size=10)

ALIGN_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_RIGHT       = Alignment(horizontal="right",  vertical="center")

BORDER_THIN       = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

PAYMENT_METHOD    = "A VISTA"
FORMAT_CURRENCY   = 'R$ #,##0.00;(R$ #,##0.00)'
FORMAT_DATE       = 'DD/MM/YYYY'


def _fill(color: str) -> PatternFill:
    """Returns a solid PatternFill for the given hex color string."""
    return PatternFill("solid", start_color=color, fgColor=color)


# ─────────────────────────────────────────────
# Function 1 — generate_all_spreadsheets
# ─────────────────────────────────────────────

def generate_all_spreadsheets(
    people: list[str],
    reference_month: str,
    due_date: str,
    output_dir: Path,
) -> list[Path]:
    """
    Entry point. Validates inputs, ensures the output directory
    exists, and generates one spreadsheet per person.

    Errors generating a single person's file are logged and skipped
    so that the remaining files are still created successfully.

    Args:
        people:          List of cardholder full names.
        reference_month: Month label used as the sheet name.
                         Example: "MAIO 2025"
        due_date:        Invoice due date string shown in the header.
                         Example: "15/06/2025"
        output_dir:      Directory where .xlsx files will be saved.

    Returns:
        List of Path objects for each successfully created file.

    Raises:
        ValueError: If people list is empty or reference_month is blank.
    """
    if not people:
        raise ValueError(
            "The 'people' list is empty. "
            "Add at least one name to the PEOPLE constant before running."
        )

    if not reference_month or not reference_month.strip():
        raise ValueError(
            "reference_month cannot be empty. "
            "Example: 'MAIO 2025'"
        )

    if not due_date or not due_date.strip():
        raise ValueError(
            "due_date cannot be empty. "
            "Example: '15/06/2025'"
        )

    _ensure_output_directory(output_dir)

    created_files: list[Path] = []

    logger.info(
        f"Generating spreadsheets for {len(people)} person(s) "
        f"| Month: {reference_month} | Due: {due_date}"
    )

    for person_name in people:
        if not person_name or not person_name.strip():
            logger.warning("Skipping empty person name in PEOPLE list.")
            continue
        try:
            file_path = generate_person_spreadsheet(
                person_name=person_name.strip(),
                reference_month=reference_month.strip(),
                due_date=due_date.strip(),
                output_dir=output_dir,
            )
            created_files.append(file_path)
        except Exception as exc:
            logger.error(
                f"Failed to generate spreadsheet for '{person_name}': {exc}"
            )

    logger.info(
        f"Done. {len(created_files)}/{len(people)} spreadsheet(s) created "
        f"in '{output_dir}/'."
    )

    return created_files


# ─────────────────────────────────────────────
# Function 2 — generate_person_spreadsheet
# ─────────────────────────────────────────────

def generate_person_spreadsheet(
    person_name: str,
    reference_month: str,
    due_date: str,
    output_dir: Path,
) -> Path:
    """
    Creates a single .xlsx spreadsheet for one cardholder.

    The spreadsheet contains one sheet (tab) named after the
    reference month, with the following structure:
        Row 1:   Title bar — "Corporate Credit Card Control"
        Row 2:   Cardholder name
        Row 3:   Reference month and due date
        Row 4:   Column headers
        Rows 5+: Blank transaction rows (national purchases)
        SUB TOTAL row with SUM formula
        International purchases section
        SUB TOTAL row (international)
        Annuity installment row
        TOTAL row with combined SUM formula
        Signature line

    The output file is named:
        Card_{FirstName}_{LastName}_{MONTH}_{YEAR}.xlsx

    Args:
        person_name:     Full name of the cardholder.
        reference_month: Sheet name and header label. E.g. "MAIO 2025"
        due_date:        Due date string. E.g. "15/06/2025"
        output_dir:      Directory to save the file.

    Returns:
        Path to the created .xlsx file.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = reference_month

    _set_column_widths(worksheet)
    _build_header_rows(worksheet, person_name, reference_month, due_date)
    _build_column_headers(worksheet, row=4)

    # National transactions section
    national_start = 5
    national_end   = national_start + TRANSACTION_ROWS - 1
    _build_transaction_rows(worksheet, national_start, national_end)

    subtotal_national_row = national_end + 1
    _build_subtotal_row(worksheet, subtotal_national_row, national_start, national_end)

    # International transactions section
    intl_header_row = subtotal_national_row + 1
    intl_start      = intl_header_row + 1
    intl_end        = intl_start + INTL_ROWS - 1
    _build_international_section(worksheet, intl_header_row, intl_start, intl_end)

    subtotal_intl_row = intl_end + 1
    _build_subtotal_row(worksheet, subtotal_intl_row, intl_start, intl_end)

    annuity_row = subtotal_intl_row + 1
    _build_annuity_row(worksheet, annuity_row)

    total_row = annuity_row + 1
    _build_total_row(worksheet, total_row, subtotal_national_row, subtotal_intl_row, annuity_row)

    signature_row = total_row + 2
    _build_signature_row(worksheet, signature_row)

    file_path = _build_output_path(output_dir, person_name, reference_month)
    workbook.save(str(file_path))

    logger.info(f"✅ Created: {file_path.name}")
    return file_path


# ─────────────────────────────────────────────
# Function 3 — _build_header_rows
# ─────────────────────────────────────────────

def _build_header_rows(
    worksheet,
    person_name: str,
    reference_month: str,
    due_date: str,
) -> None:
    """
    Builds the first three header rows of the spreadsheet:

        Row 1 — Dark blue title bar spanning all columns:
                "Corporate Credit Card Control"

        Row 2 — Person label + cardholder name:
                "Cardholder:" | "Aline Maris"

        Row 3 — Month reference + due date:
                "Reference:" | "MAIO 2025 — Due date: 15/06/2025"

    All cells use Arial font, borders, and appropriate alignment.
    The title row uses white bold text on a dark blue background.
    Rows 2 and 3 use normal weight text on a white background.

    Args:
        worksheet:       The active openpyxl Worksheet.
        person_name:     Full name of the cardholder.
        reference_month: Month label. Example: "MAIO 2025"
        due_date:        Due date. Example: "15/06/2025"

    Returns:
        None. Modifies the worksheet in place.
    """
    last_col = get_column_letter(TOTAL_COLUMNS)

    # Row 1 — Title
    worksheet.merge_cells(f"A1:{last_col}1")
    cell = worksheet["A1"]
    cell.value     = "Corporate Credit Card Control"
    cell.font      = FONT_TITLE
    cell.fill      = _fill(COLOR_DARK_BLUE)
    cell.alignment = ALIGN_CENTER
    cell.border    = BORDER_THIN
    worksheet.row_dimensions[1].height = 22

    # Row 2 — Cardholder
    worksheet.merge_cells("A2:B2")
    worksheet["A2"].value     = "Cardholder:"
    worksheet["A2"].font      = FONT_BOLD
    worksheet["A2"].alignment = ALIGN_LEFT
    worksheet["A2"].border    = BORDER_THIN

    worksheet.merge_cells(f"C2:{last_col}2")
    worksheet["C2"].value     = person_name
    worksheet["C2"].font      = FONT_NORMAL
    worksheet["C2"].alignment = ALIGN_LEFT
    worksheet["C2"].border    = BORDER_THIN
    worksheet.row_dimensions[2].height = 18

    # Row 3 — Reference month and due date
    worksheet.merge_cells("A3:B3")
    worksheet["A3"].value     = "Reference:"
    worksheet["A3"].font      = FONT_BOLD
    worksheet["A3"].alignment = ALIGN_LEFT
    worksheet["A3"].border    = BORDER_THIN

    worksheet.merge_cells(f"C3:{last_col}3")
    worksheet["C3"].value     = f"{reference_month} — Due date: {due_date}"
    worksheet["C3"].font      = FONT_NORMAL
    worksheet["C3"].alignment = ALIGN_LEFT
    worksheet["C3"].border    = BORDER_THIN
    worksheet.row_dimensions[3].height = 18


# ─────────────────────────────────────────────
# Function 4 — _build_transaction_rows
# ─────────────────────────────────────────────

def _build_transaction_rows(
    worksheet,
    start_row: int,
    end_row: int,
) -> None:
    """
    Creates blank formatted rows for transaction data entry
    between start_row and end_row (inclusive).

    Column-specific formatting applied:
        COL 1  (DATA)          → date format DD/MM/YYYY, centered
        COL 2  (NF/CUPOM)      → text, centered
        COL 3  (FORNECEDOR)    → text, left-aligned
        COL 4  (DESCRIÇÃO)     → text, left-aligned
        COL 5  (VALOR)         → currency R$ #,##0.00, right-aligned
        COL 6  (FORMA DE PGTO) → pre-filled "A VISTA", centered
        COL 7  (PARCELAS)      → text, centered
        COL 8  (PROJETO)       → text, left-aligned
        COL 9  (REEMBOLSÁVEL)  → text, centered
        COL 10 (GESTOR)        → text, left-aligned
        COL 11 (OBSERVAÇÃO)    → text, left-aligned

    Each row has a height of 18pt and a thin border on all cells.

    Args:
        worksheet: The active openpyxl Worksheet.
        start_row: First data row number (inclusive).
        end_row:   Last data row number (inclusive).

    Returns:
        None. Modifies the worksheet in place.
    """
    column_config = {
        1:  (FORMAT_DATE,     ALIGN_CENTER, None),
        2:  (None,            ALIGN_CENTER, None),
        3:  (None,            ALIGN_LEFT,   None),
        4:  (None,            ALIGN_LEFT,   None),
        5:  (FORMAT_CURRENCY, ALIGN_RIGHT,  None),
        6:  (None,            ALIGN_CENTER, PAYMENT_METHOD),
        7:  (None,            ALIGN_CENTER, None),
        8:  (None,            ALIGN_LEFT,   None),
        9:  (None,            ALIGN_CENTER, None),
        10: (None,            ALIGN_LEFT,   None),
        11: (None,            ALIGN_LEFT,   None),
    }

    for row in range(start_row, end_row + 1):
        worksheet.row_dimensions[row].height = 18

        for col_idx in range(1, TOTAL_COLUMNS + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.font   = FONT_NORMAL
            cell.border = BORDER_THIN

            num_fmt, alignment, default_value = column_config.get(
                col_idx, (None, ALIGN_LEFT, None)
            )
            cell.alignment = alignment

            if num_fmt:
                cell.number_format = num_fmt

            if default_value is not None:
                cell.value = default_value


# ─────────────────────────────────────────────
# Function 5 — _build_summary_rows
# ─────────────────────────────────────────────

def _build_summary_rows(
    worksheet,
    subtotal_national_row: int,
    subtotal_intl_row: int,
    annuity_row: int,
    total_row: int,
    signature_row: int,
) -> None:
    """
    Builds all financial summary rows at the bottom of the sheet:

        SUB TOTAL (national)      → SUM formula over national rows
        SUB TOTAL (international) → SUM formula over intl rows
        Annuity installment       → Fixed amount (ANNUITY_AMOUNT)
        TOTAL                     → Sum of the three rows above
        Signature line            → "Sign. Card Manager ___________"

    This function is the orchestrator for the summary section.
    Each individual row is built by its own private helper to keep
    each function focused and under the complexity limit.

    Color coding:
        SUB TOTAL rows  → Light blue background
        Annuity row     → White background
        TOTAL row       → Dark blue background with white text

    Args:
        worksheet:             The active openpyxl Worksheet.
        subtotal_national_row: Row index of the national SUB TOTAL.
        subtotal_intl_row:     Row index of the international SUB TOTAL.
        annuity_row:           Row index of the annuity row.
        total_row:             Row index of the TOTAL row.
        signature_row:         Row index of the signature line.

    Returns:
        None. Modifies the worksheet in place.
    """
    last_col = get_column_letter(TOTAL_COLUMNS)

    # Style the SUB TOTAL label cells (already created by _build_subtotal_row)
    for row in [subtotal_national_row, subtotal_intl_row]:
        worksheet.row_dimensions[row].height = 18
        for col_idx in range(1, TOTAL_COLUMNS + 1):
            cell = worksheet.cell(row=row, column=col_idx)
            cell.fill   = _fill(COLOR_LIGHT_BLUE)
            cell.border = BORDER_THIN

    # Style the TOTAL row
    worksheet.row_dimensions[total_row].height = 22
    for col_idx in range(1, TOTAL_COLUMNS + 1):
        cell = worksheet.cell(row=total_row, column=col_idx)
        cell.fill   = _fill(COLOR_DARK_BLUE)
        cell.border = BORDER_THIN

    # Signature line
    worksheet.merge_cells(f"A{signature_row}:{last_col}{signature_row}")
    sig_cell = worksheet[f"A{signature_row}"]
    sig_cell.value     = "Sign. Card Manager ___________________________________________________"
    sig_cell.font      = FONT_NORMAL
    sig_cell.alignment = ALIGN_LEFT
    worksheet.row_dimensions[signature_row].height = 18


# ─────────────────────────────────────────────
# Private helpers
# ─────────────────────────────────────────────

def _set_column_widths(worksheet) -> None:
    """Sets the width of each column based on the COLUMNS definition."""
    for index, col_def in enumerate(COLUMNS, start=1):
        worksheet.column_dimensions[get_column_letter(index)].width = col_def.width


def _build_column_headers(worksheet, row: int) -> None:
    """
    Builds the column header row with blue background and white bold text.
    Each column header matches the COLUMNS definition.
    """
    worksheet.row_dimensions[row].height = 32
    for col_idx, col_def in enumerate(COLUMNS, start=1):
        cell = worksheet.cell(row=row, column=col_idx, value=col_def.header)
        cell.font      = FONT_HEADER_COL
        cell.fill      = _fill(COLOR_MED_BLUE)
        cell.alignment = ALIGN_CENTER
        cell.border    = BORDER_THIN


def _build_subtotal_row(
    worksheet,
    subtotal_row: int,
    data_start: int,
    data_end: int,
) -> None:
    """
    Builds a SUB TOTAL row with a SUM formula over the given data rows.
    Label spans columns A–D; formula is placed in column E (VALOR).
    """
    last_col = get_column_letter(TOTAL_COLUMNS)
    worksheet.row_dimensions[subtotal_row].height = 18
    worksheet.merge_cells(f"A{subtotal_row}:D{subtotal_row}")

    label_cell = worksheet[f"A{subtotal_row}"]
    label_cell.value     = "SUB TOTAL"
    label_cell.font      = FONT_BOLD
    label_cell.fill      = _fill(COLOR_LIGHT_BLUE)
    label_cell.alignment = ALIGN_RIGHT
    label_cell.border    = BORDER_THIN

    amount_cell = worksheet[f"E{subtotal_row}"]
    amount_cell.value         = f"=SUM(E{data_start}:E{data_end})"
    amount_cell.font          = FONT_BOLD
    amount_cell.fill          = _fill(COLOR_LIGHT_BLUE)
    amount_cell.number_format = FORMAT_CURRENCY
    amount_cell.alignment     = ALIGN_RIGHT
    amount_cell.border        = BORDER_THIN

    for col_idx in [2, 3, 4] + list(range(6, TOTAL_COLUMNS + 1)):
        cell = worksheet.cell(row=subtotal_row, column=col_idx)
        cell.fill   = _fill(COLOR_LIGHT_BLUE)
        cell.border = BORDER_THIN


def _build_international_section(
    worksheet,
    header_row: int,
    data_start: int,
    data_end: int,
) -> None:
    """
    Builds the INTERNATIONAL PURCHASES header and its blank data rows.
    Header spans all columns with a light blue background.
    """
    last_col = get_column_letter(TOTAL_COLUMNS)
    worksheet.row_dimensions[header_row].height = 18
    worksheet.merge_cells(f"A{header_row}:{last_col}{header_row}")

    header_cell = worksheet[f"A{header_row}"]
    header_cell.value     = "INTERNATIONAL PURCHASES"
    header_cell.font      = FONT_BOLD
    header_cell.fill      = _fill(COLOR_LIGHT_BLUE)
    header_cell.alignment = ALIGN_CENTER
    header_cell.border    = BORDER_THIN

    _build_transaction_rows(worksheet, data_start, data_end)


def _build_annuity_row(worksheet, row: int) -> None:
    """
    Builds the annuity installment row with a fixed amount.
    Label spans columns A–D; amount is in column E.
    """
    worksheet.row_dimensions[row].height = 18
    worksheet.merge_cells(f"A{row}:D{row}")

    label_cell = worksheet[f"A{row}"]
    label_cell.value     = "Annuity installment:"
    label_cell.font      = FONT_BOLD
    label_cell.alignment = ALIGN_RIGHT
    label_cell.border    = BORDER_THIN

    amount_cell = worksheet[f"E{row}"]
    amount_cell.value         = ANNUITY_AMOUNT
    amount_cell.font          = FONT_NORMAL
    amount_cell.number_format = FORMAT_CURRENCY
    amount_cell.alignment     = ALIGN_RIGHT
    amount_cell.border        = BORDER_THIN

    for col_idx in [2, 3, 4] + list(range(6, TOTAL_COLUMNS + 1)):
        worksheet.cell(row=row, column=col_idx).border = BORDER_THIN


def _build_total_row(
    worksheet,
    total_row: int,
    subtotal_national: int,
    subtotal_intl: int,
    annuity_row: int,
) -> None:
    """
    Builds the TOTAL row with a dark blue background and white text.
    Formula sums: national subtotal + international subtotal + annuity.
    """
    last_col = get_column_letter(TOTAL_COLUMNS)
    worksheet.row_dimensions[total_row].height = 22
    worksheet.merge_cells(f"A{total_row}:D{total_row}")

    label_cell = worksheet[f"A{total_row}"]
    label_cell.value     = "TOTAL"
    label_cell.font      = FONT_TOTAL
    label_cell.fill      = _fill(COLOR_DARK_BLUE)
    label_cell.alignment = ALIGN_RIGHT
    label_cell.border    = BORDER_THIN

    amount_cell = worksheet[f"E{total_row}"]
    amount_cell.value         = f"=E{subtotal_national}+E{subtotal_intl}+E{annuity_row}"
    amount_cell.font          = FONT_TOTAL
    amount_cell.fill          = _fill(COLOR_DARK_BLUE)
    amount_cell.number_format = FORMAT_CURRENCY
    amount_cell.alignment     = ALIGN_RIGHT
    amount_cell.border        = BORDER_THIN

    for col_idx in [2, 3, 4] + list(range(6, TOTAL_COLUMNS + 1)):
        cell = worksheet.cell(row=total_row, column=col_idx)
        cell.fill   = _fill(COLOR_DARK_BLUE)
        cell.border = BORDER_THIN


def _build_signature_row(worksheet, row: int) -> None:
    """
    Builds the signature line at the bottom of the spreadsheet.
    Spans all columns with left-aligned normal text.
    """
    last_col = get_column_letter(TOTAL_COLUMNS)
    worksheet.merge_cells(f"A{row}:{last_col}{row}")
    cell = worksheet[f"A{row}"]
    cell.value     = "Sign. Card Manager ___________________________________________________"
    cell.font      = FONT_NORMAL
    cell.alignment = ALIGN_LEFT
    worksheet.row_dimensions[row].height = 18


def _ensure_output_directory(output_dir: Path) -> None:
    """
    Creates the output directory if it does not already exist.
    Raises ValueError if the path exists but is not a directory.
    """
    if output_dir.exists() and not output_dir.is_dir():
        raise ValueError(
            f"Output path '{output_dir}' exists but is not a directory. "
            "Please provide a valid directory path."
        )
    output_dir.mkdir(parents=True, exist_ok=True)
    logger.debug(f"Output directory ready: '{output_dir}'")


def _build_output_path(output_dir: Path, person_name: str, reference_month: str) -> Path:
    """
    Builds the output file path from the person name and month.

    Naming convention:
        Card_{FirstName}_{LastName}_{MONTH}_{YEAR}.xlsx
        Example: Card_Aline_Maris_MAIO_2025.xlsx

    Special characters and spaces are replaced with underscores.
    """
    safe_name  = person_name.replace(" ", "_").replace("/", "_")
    safe_month = reference_month.replace(" ", "_").replace("/", "_")
    file_name  = f"Card_{safe_name}_{safe_month}.xlsx"
    return output_dir / file_name


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

if __name__ == "__main__":
    try:
        created = generate_all_spreadsheets(
            people=PEOPLE,
            reference_month=REFERENCE_MONTH,
            due_date=DUE_DATE,
            output_dir=OUTPUT_DIR,
        )
        print(f"\n{len(created)} spreadsheet(s) created in '{OUTPUT_DIR}/':")
        for path in created:
            print(f"  • {path.name}")
    except ValueError as exc:
        logger.error(f"Configuration error: {exc}")
        sys.exit(1)