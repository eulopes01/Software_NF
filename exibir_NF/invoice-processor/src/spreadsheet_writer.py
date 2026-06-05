"""
spreadsheet_writer.py — Phase 3: Spreadsheet Writer
-----------------------------------------------------
Responsible for opening the person's .xlsx spreadsheet,
locating the correct month sheet, inserting extracted
transactions, formatting cells, and saving the file.

Each function has a single responsibility and at least 25 lines,
following Clean Architecture, SOLID, and Secure by Design principles.

Dependencies:
    pip install openpyxl

Usage:
    from spreadsheet_writer import write_transactions

    write_transactions(
        spreadsheet_path="spreadsheets/card_john.xlsx",
        sheet_name="MAIO 2025",
        transactions=transactions,
    )
"""

import logging
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pdf_reader import Transaction

# ─────────────────────────────────────────────
# Logging setup
# ─────────────────────────────────────────────

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────

ALLOWED_EXTENSION      = ".xlsx"
MAX_FILE_SIZE_MB       = 20
MAX_TRANSACTIONS       = 500
PAYMENT_METHOD         = "A VISTA"

# Column indexes (1-based) matching the spreadsheet template
COL_DATE               = 1   # DATA
COL_NF                 = 2   # NF/CUPOM         — left blank (manual)
COL_SUPPLIER           = 3   # FORNECEDOR
COL_DESCRIPTION        = 4   # DESCRIÇÃO
COL_AMOUNT             = 5   # VALOR
COL_PAYMENT_METHOD     = 6   # FORMA DE PGTO
COL_INSTALLMENTS       = 7   # PARCELAS         — left blank (manual)
COL_PROJECT            = 8   # PROJETO          — left blank (manual)
COL_REIMBURSABLE       = 9   # REEMBOLSÁVEL     — left blank (manual)
COL_MANAGER            = 10  # GESTOR           — left blank (manual)
COL_OBSERVATION        = 11  # OBSERVAÇÃO       — left blank (manual)
TOTAL_COLUMNS          = 11

# Row where the data entry starts (after header rows 1–4)
DATA_START_ROW         = 5

# Markers used to find the first empty data row in the sheet
END_SECTION_MARKERS    = [
    "sub total",
    "total",
    "compras internacionais",
    "parcela de anuidade",
    "ass. gestor",
]

# Styles
FONT_NORMAL            = Font(name="Arial", size=10)
FONT_BOLD              = Font(name="Arial", bold=True, size=10)
ALIGNMENT_CENTER       = Alignment(horizontal="center", vertical="center")
ALIGNMENT_LEFT         = Alignment(horizontal="left",   vertical="center")
ALIGNMENT_RIGHT        = Alignment(horizontal="right",  vertical="center")
BORDER_THIN            = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
FILL_DATA_ROW          = PatternFill("solid", start_color="FFFFFF", fgColor="FFFFFF")


# ─────────────────────────────────────────────
# Function 1 — write_transactions
# ─────────────────────────────────────────────

def write_transactions(
    spreadsheet_path: str,
    sheet_name: str,
    transactions: list[Transaction],
) -> int:
    """
    Main entry point for the spreadsheet writing pipeline.

    Validates inputs, opens the workbook, locates the correct sheet,
    unmerges header cells if needed, inserts all transactions,
    formats the cells, and saves the file safely.

    Args:
        spreadsheet_path: Path to the person's .xlsx file.
        sheet_name: Name of the sheet (tab) for the target month.
        transactions: List of normalized Transaction objects.

    Returns:
        The number of transactions successfully written.

    Raises:
        FileNotFoundError: If the spreadsheet file does not exist.
        ValueError: If inputs are invalid or the sheet is not found.
    """
    path = Path(spreadsheet_path)

    logger.info(
        f"Starting spreadsheet write | File: {path.name} "
        f"| Sheet: {sheet_name} | Transactions: {len(transactions)}"
    )

    _validate_spreadsheet_inputs(path, sheet_name, transactions)

    workbook  = open_spreadsheet(spreadsheet_path)
    worksheet = locate_sheet(workbook, sheet_name)

    _unmerge_data_area(worksheet)

    first_empty_row = _find_first_empty_data_row(worksheet)
    written_count   = insert_rows(worksheet, transactions, first_empty_row)
    format_cells(worksheet, first_empty_row, written_count)
    save_spreadsheet(workbook, spreadsheet_path)

    logger.info(
        f"Write complete | {written_count} transactions inserted "
        f"| File: {path.name} | Sheet: {sheet_name}"
    )

    return written_count


# ─────────────────────────────────────────────
# Function 2 — open_spreadsheet
# ─────────────────────────────────────────────

def open_spreadsheet(spreadsheet_path: str):
    """
    Opens and returns an openpyxl Workbook object from the given path.

    Validates that the file exists, is a valid .xlsx file, and is
    within the allowed size limit before attempting to open it.
    Uses data_only=False to preserve formulas (SUM formulas in
    SUB TOTAL and TOTAL rows must remain intact).

    Args:
        spreadsheet_path: Absolute or relative path to the .xlsx file.

    Returns:
        An openpyxl Workbook object ready for reading and writing.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the file cannot be opened or is corrupted.
    """
    path = Path(spreadsheet_path)

    if not path.exists():
        raise FileNotFoundError(
            f"Spreadsheet not found: '{path}'. "
            "Make sure the file was generated using generate_spreadsheet.py."
        )

    if not path.is_file():
        raise ValueError(f"Path is not a file: '{path}'")

    if path.suffix.lower() != ALLOWED_EXTENSION:
        raise ValueError(
            f"Invalid file type '{path.suffix}'. "
            f"Only '{ALLOWED_EXTENSION}' files are supported."
        )

    file_size_mb = path.stat().st_size / (1024 * 1024)
    if file_size_mb > MAX_FILE_SIZE_MB:
        raise ValueError(
            f"File '{path.name}' exceeds the maximum allowed size "
            f"({file_size_mb:.1f} MB > {MAX_FILE_SIZE_MB} MB)."
        )

    try:
        workbook = load_workbook(str(path), data_only=False)
        logger.debug(
            f"Spreadsheet opened: {path.name} "
            f"| Sheets: {workbook.sheetnames}"
        )
        return workbook
    except Exception as exc:
        raise ValueError(
            f"Failed to open spreadsheet '{path.name}': {exc}"
        ) from exc


# ─────────────────────────────────────────────
# Function 3 — locate_sheet
# ─────────────────────────────────────────────

def locate_sheet(workbook, sheet_name: str):
    """
    Finds and returns the worksheet matching the given sheet name
    inside the provided workbook.

    Matching is case-insensitive and strips surrounding whitespace
    to handle common naming inconsistencies (e.g. "maio 2025"
    matches "MAIO 2025"). If an exact match is not found, a
    normalized comparison is attempted.

    Args:
        workbook: An open openpyxl Workbook object.
        sheet_name: The target sheet name. Example: "MAIO 2025".

    Returns:
        The openpyxl Worksheet object for the matched sheet.

    Raises:
        ValueError: If no matching sheet is found in the workbook.
    """
    if not sheet_name or not isinstance(sheet_name, str):
        raise ValueError(
            "sheet_name must be a non-empty string. "
            f"Received: {sheet_name!r}"
        )

    normalized_target = sheet_name.strip().upper()

    if sheet_name in workbook.sheetnames:
        logger.debug(f"Sheet found (exact match): '{sheet_name}'")
        return workbook[sheet_name]

    for name in workbook.sheetnames:
        if name.strip().upper() == normalized_target:
            logger.debug(f"Sheet found (normalized match): '{name}'")
            return workbook[name]

    available = ", ".join(f"'{s}'" for s in workbook.sheetnames)
    raise ValueError(
        f"Sheet '{sheet_name}' not found in the workbook. "
        f"Available sheets: [{available}]. "
        "Make sure the spreadsheet was generated with the correct month name."
    )


# ─────────────────────────────────────────────
# Function 4 — insert_rows
# ─────────────────────────────────────────────

def insert_rows(
    worksheet,
    transactions: list[Transaction],
    start_row: int,
) -> int:
    """
    Writes each Transaction as a new row in the worksheet,
    starting at the given row number.

    Columns written automatically:
        COL_DATE           → transaction.date       (DD/MM)
        COL_SUPPLIER       → transaction.description (first 30 chars)
        COL_DESCRIPTION    → transaction.description (full, up to 100 chars)
        COL_AMOUNT         → transaction.amount (negative shown as text)
        COL_PAYMENT_METHOD → "A VISTA" (always)

    Columns left blank for manual entry:
        COL_NF, COL_INSTALLMENTS, COL_PROJECT,
        COL_REIMBURSABLE, COL_MANAGER, COL_OBSERVATION

    Args:
        worksheet: An open openpyxl Worksheet object.
        transactions: List of normalized Transaction objects.
        start_row: The row number where insertion begins.

    Returns:
        The number of rows successfully written.
    """
    if not transactions:
        logger.warning("insert_rows called with empty transaction list. Nothing written.")
        return 0

    if start_row < DATA_START_ROW:
        raise ValueError(
            f"start_row ({start_row}) is before the data area "
            f"(minimum: {DATA_START_ROW}). "
            "This would overwrite header rows."
        )

    written = 0

    for index, transaction in enumerate(transactions):
        current_row = start_row + index
        supplier    = transaction.description[:30] if transaction.description else ""

        # Format negative amounts as text "(-value)" so they don't subtract
        if transaction.amount < 0:
            abs_val   = abs(transaction.amount)
            formatted = f"{abs_val:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            amount_value = f"(-{formatted})"
        else:
            amount_value = transaction.amount

        # Write all columns safely — skips MergedCell objects
        column_values = [
            (COL_DATE,           transaction.date),
            (COL_NF,             None),
            (COL_SUPPLIER,       supplier),
            (COL_DESCRIPTION,    transaction.description),
            (COL_AMOUNT,         amount_value),
            (COL_PAYMENT_METHOD, PAYMENT_METHOD),
            (COL_INSTALLMENTS,   None),
            (COL_PROJECT,        None),
            (COL_REIMBURSABLE,   None),
            (COL_MANAGER,        None),
            (COL_OBSERVATION,    None),
        ]

        for col, val in column_values:
            cell = worksheet.cell(row=current_row, column=col)
            if cell.__class__.__name__ != "MergedCell":
                cell.value = val

        written += 1
        logger.debug(
            f"Row {current_row} written: {transaction.date} | "
            f"{transaction.description[:40]} | R$ {transaction.amount:.2f}"
        )

    logger.info(f"insert_rows complete: {written} rows written starting at row {start_row}.")
    return written


# ─────────────────────────────────────────────
# Function 5 — format_cells
# ─────────────────────────────────────────────

def format_cells(worksheet, start_row: int, row_count: int) -> None:
    """
    Applies consistent visual formatting to all newly inserted
    transaction rows in the worksheet.

    Formatting applied per column:
        COL_DATE           → centered, Arial 10
        COL_NF             → centered, Arial 10
        COL_SUPPLIER       → left-aligned, Arial 10
        COL_DESCRIPTION    → left-aligned, Arial 10
        COL_AMOUNT         → right-aligned, Arial 10, BRL currency format
        COL_PAYMENT_METHOD → centered, Arial 10
        All others         → left-aligned, Arial 10

    A thin border is applied to all cells in each written row.
    Row height is set to 18pt for consistent visual spacing.
    Skips MergedCell objects to avoid AttributeError.

    Args:
        worksheet: An open openpyxl Worksheet object.
        start_row: The first row that was written by insert_rows().
        row_count: The number of rows that were written.

    Returns:
        None
    """
    if row_count <= 0:
        logger.debug("format_cells skipped: row_count is 0.")
        return

    currency_format = "R$ #,##0.00"

    column_styles = {
        COL_DATE:           (ALIGNMENT_CENTER, FONT_NORMAL, None),
        COL_NF:             (ALIGNMENT_CENTER, FONT_NORMAL, None),
        COL_SUPPLIER:       (ALIGNMENT_LEFT,   FONT_NORMAL, None),
        COL_DESCRIPTION:    (ALIGNMENT_LEFT,   FONT_NORMAL, None),
        COL_AMOUNT:         (ALIGNMENT_RIGHT,  FONT_NORMAL, currency_format),
        COL_PAYMENT_METHOD: (ALIGNMENT_CENTER, FONT_NORMAL, None),
        COL_INSTALLMENTS:   (ALIGNMENT_CENTER, FONT_NORMAL, None),
        COL_PROJECT:        (ALIGNMENT_LEFT,   FONT_NORMAL, None),
        COL_REIMBURSABLE:   (ALIGNMENT_CENTER, FONT_NORMAL, None),
        COL_MANAGER:        (ALIGNMENT_LEFT,   FONT_NORMAL, None),
        COL_OBSERVATION:    (ALIGNMENT_LEFT,   FONT_NORMAL, None),
    }

    for row_offset in range(row_count):
        current_row = start_row + row_offset
        worksheet.row_dimensions[current_row].height = 18

        for col_index in range(1, TOTAL_COLUMNS + 1):
            cell = worksheet.cell(row=current_row, column=col_index)

            # Skip merged cells — they cannot be styled directly
            if cell.__class__.__name__ == "MergedCell":
                continue

            alignment, font, number_format = column_styles.get(
                col_index, (ALIGNMENT_LEFT, FONT_NORMAL, None)
            )
            cell.alignment = alignment
            cell.font      = font
            cell.border    = BORDER_THIN

            # Only apply currency format to numeric cells
            if number_format and not isinstance(cell.value, str):
                cell.number_format = number_format

    logger.debug(
        f"format_cells applied to rows {start_row}–{start_row + row_count - 1}."
    )


# ─────────────────────────────────────────────
# Function 6 — save_spreadsheet
# ─────────────────────────────────────────────

def save_spreadsheet(workbook, spreadsheet_path: str) -> None:
    """
    Saves the modified workbook back to the original file path.

    Performs a safe save: writes to a temporary file first, then
    replaces the original only on success. This prevents data loss
    if the save operation is interrupted (e.g. disk full, crash).

    The temporary file is placed in the same directory as the
    original to ensure atomic rename works across filesystems.

    Args:
        workbook: The modified openpyxl Workbook object.
        spreadsheet_path: The destination file path (original file).

    Raises:
        ValueError: If saving fails for any reason.
    """
    path      = Path(spreadsheet_path)
    temp_path = path.with_suffix(".tmp.xlsx")

    logger.debug(f"Saving spreadsheet to temp file: {temp_path.name}")

    try:
        workbook.save(str(temp_path))
    except Exception as exc:
        if temp_path.exists():
            temp_path.unlink()
        raise ValueError(
            f"Failed to save spreadsheet '{path.name}': {exc}"
        ) from exc

    try:
        temp_path.replace(path)
        logger.info(f"Spreadsheet saved successfully: {path.name}")
    except Exception as exc:
        raise ValueError(
            f"Failed to replace original file '{path.name}' "
            f"with saved version: {exc}"
        ) from exc


# ─────────────────────────────────────────────
# Private helpers
# ─────────────────────────────────────────────

def _unmerge_data_area(worksheet) -> None:
    """
    Unmerges any merged cell ranges that overlap with the data
    entry area (rows DATA_START_ROW and beyond).

    This prevents 'MergedCell object attribute value is read-only'
    errors when writing transaction data into rows that were
    previously part of a merged range.

    Only ranges that start at or after DATA_START_ROW are unmerged
    to preserve the header formatting (rows 1–4).

    Args:
        worksheet: The active openpyxl Worksheet object.

    Returns:
        None. Modifies the worksheet in place.
    """
    ranges_to_unmerge = [
        str(r) for r in worksheet.merged_cells.ranges
        if r.min_row >= DATA_START_ROW
    ]

    for cell_range in ranges_to_unmerge:
        try:
            worksheet.unmerge_cells(cell_range)
            logger.debug(f"Unmerged cells: {cell_range}")
        except Exception as exc:
            logger.warning(f"Could not unmerge {cell_range}: {exc}")

    if ranges_to_unmerge:
        logger.debug(f"Unmerged {len(ranges_to_unmerge)} range(s) in data area.")


def _validate_spreadsheet_inputs(
    path: Path,
    sheet_name: str,
    transactions: list,
) -> None:
    """
    Validates all inputs before the write pipeline begins.

    Raises:
        FileNotFoundError: If the spreadsheet file does not exist.
        ValueError: If sheet_name is empty or transactions exceed limit.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Spreadsheet not found: '{path}'. "
            "Run generate_spreadsheet.py to create it first."
        )

    if not sheet_name or not sheet_name.strip():
        raise ValueError("sheet_name cannot be empty.")

    if not isinstance(transactions, list):
        raise ValueError(
            f"transactions must be a list. Received: {type(transactions)}"
        )

    if len(transactions) > MAX_TRANSACTIONS:
        raise ValueError(
            f"Too many transactions ({len(transactions)}). "
            f"Maximum allowed per sheet: {MAX_TRANSACTIONS}."
        )

    logger.debug(
        f"Input validation passed: file='{path.name}' "
        f"sheet='{sheet_name}' transactions={len(transactions)}"
    )


def _create_backup(path: Path) -> None:
    """
    Creates a timestamped backup copy of the spreadsheet before
    any modifications are made.

    Backup file is placed in the same directory as the original.
    If the backup fails, a warning is logged but execution continues.
    """
    timestamp   = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"{path.stem}.{timestamp}.backup{path.suffix}"
    backup_path = path.parent / backup_name

    try:
        shutil.copy2(str(path), str(backup_path))
        logger.info(f"Backup created: {backup_name}")
    except Exception as exc:
        logger.warning(
            f"Could not create backup for '{path.name}': {exc}. "
            "Proceeding without backup."
        )


def _find_first_empty_data_row(worksheet) -> int:
    """
    Scans the worksheet from DATA_START_ROW downward to find
    the first row that is empty in the DATE column (column 1)
    and does not contain a section marker (subtotal, total, etc).

    This ensures new transactions are always appended after
    existing ones, and never overwrite summary rows.

    Returns:
        The row number of the first available empty data row.
    """
    row = DATA_START_ROW

    while True:
        cell      = worksheet.cell(row=row, column=COL_DATE)
        cell_value = None if cell.__class__.__name__ == "MergedCell" else cell.value

        if cell_value is None or str(cell_value).strip() == "":
            desc_cell  = worksheet.cell(row=row, column=COL_DESCRIPTION)
            desc_value = None if desc_cell.__class__.__name__ == "MergedCell" else desc_cell.value
            desc_str   = str(desc_value).strip().lower() if desc_value else ""

            if not any(marker in desc_str for marker in END_SECTION_MARKERS):
                logger.debug(f"First empty data row found: {row}")
                return row

        row += 1

        if row > DATA_START_ROW + MAX_TRANSACTIONS:
            logger.warning(
                f"Could not find empty row after scanning {MAX_TRANSACTIONS} rows. "
                f"Defaulting to row {DATA_START_ROW}."
            )
            return DATA_START_ROW