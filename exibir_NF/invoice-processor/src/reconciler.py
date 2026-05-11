"""
reconciler.py — Phase 4: Invoice Reconciler
---------------------------------------------
Responsible for reading the invoice total from the PDF,
reading the computed total from the spreadsheet, comparing
both values, reporting the result (match or divergence),
and writing a structured log entry for audit purposes.

Each function has a single responsibility and at least 25 lines,
following Clean Architecture, SOLID, and Secure by Design principles.

Dependencies:
    pip install openpyxl

Usage:
    from reconciler import reconcile

    result = reconcile(
        person_name="Aline Maris",
        sheet_name="MAIO 2025",
        spreadsheet_path="spreadsheets/card_Aline_Maris_MAIO_2025.xlsx",
        pdf_total=1337.61,
    )
"""

import logging
import json
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

# ─────────────────────────────────────────────
# Logging setup
# ─────────────────────────────────────────────

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────
# Constants
# ─────────────────────────────────────────────

# Tolerance in BRL to handle floating point rounding differences
# Example: 1337.61 vs 1337.6099999 → considered a match
RECONCILE_TOLERANCE     = 0.01

# Column index of VALOR (amount) in the spreadsheet
COL_AMOUNT              = 5

# Markers that identify the SUB TOTAL row in the spreadsheet
SUBTOTAL_MARKERS        = ["sub total", "subtotal"]

# Default log directory and file
DEFAULT_LOG_DIR         = Path("logs")
DEFAULT_LOG_FILE        = DEFAULT_LOG_DIR / "processing.log"

# Result status constants
STATUS_MATCH            = "MATCH"
STATUS_DIVERGENCE       = "DIVERGENCE"
STATUS_ERROR            = "ERROR"

# Maximum rows to scan when searching for the subtotal row
MAX_SCAN_ROWS           = 600

# Allowed spreadsheet extension
ALLOWED_EXTENSION       = ".xlsx"


# ─────────────────────────────────────────────
# Data model
# ─────────────────────────────────────────────

@dataclass
class ReconciliationResult:
    """
    Structured result of a single invoice reconciliation operation.

    Attributes:
        person_name:       Full name of the cardholder.
        sheet_name:        Name of the worksheet tab processed.
        spreadsheet_path:  Path to the .xlsx file.
        pdf_total:         Total value extracted from the PDF invoice.
        spreadsheet_total: Total value computed from the spreadsheet.
        difference:        Absolute difference between both totals.
        status:            "MATCH", "DIVERGENCE", or "ERROR".
        message:           Human-readable summary of the result.
        timestamp:         ISO-8601 timestamp of when reconciliation ran.
    """
    person_name:       str
    sheet_name:        str
    spreadsheet_path:  str
    pdf_total:         float
    spreadsheet_total: float
    difference:        float
    status:            str
    message:           str
    timestamp:         str


# ─────────────────────────────────────────────
# Function 1 — reconcile
# ─────────────────────────────────────────────

def reconcile(
    person_name: str,
    sheet_name: str,
    spreadsheet_path: str,
    pdf_total: float,
) -> ReconciliationResult:
    """
    Main entry point for the reconciliation pipeline.

    Reads the spreadsheet total, compares it against the PDF total,
    builds a structured result, reports it to the console, and
    writes a log entry for audit purposes.

    Args:
        person_name:      Full name of the cardholder. Used in reporting.
        sheet_name:       Name of the worksheet tab to reconcile.
        spreadsheet_path: Path to the person's .xlsx file.
        pdf_total:        The total invoice value extracted from the PDF.

    Returns:
        A ReconciliationResult dataclass with all comparison details.
    """
    timestamp = datetime.now().isoformat(timespec="seconds")

    logger.info(
        f"Starting reconciliation | Person: {person_name} "
        f"| Sheet: {sheet_name} | PDF total: R$ {pdf_total:.2f}"
    )

    _validate_reconcile_inputs(person_name, sheet_name, spreadsheet_path, pdf_total)

    spreadsheet_total = read_spreadsheet_total(spreadsheet_path, sheet_name)
    difference = read_difference(pdf_total, spreadsheet_total)
    status, message = compare_totals(person_name, pdf_total, spreadsheet_total, difference)

    result = ReconciliationResult(
        person_name=person_name,
        sheet_name=sheet_name,
        spreadsheet_path=str(spreadsheet_path),
        pdf_total=round(pdf_total, 2),
        spreadsheet_total=round(spreadsheet_total, 2),
        difference=round(difference, 2),
        status=status,
        message=message,
        timestamp=timestamp,
    )

    report_result(result)
    write_log(result)

    return result


# ─────────────────────────────────────────────
# Function 2 — read_pdf_total (wrapper/validator)
# ─────────────────────────────────────────────

def read_pdf_total(pdf_total_raw: float, file_name: str = "") -> float:
    """
    Validates and returns the PDF invoice total value.

    The actual PDF extraction is performed by pdf_reader.read_pdf().
    This function acts as a validation and normalization layer before
    the value enters the reconciliation pipeline, ensuring the total
    is a valid positive float and within a reasonable range for a
    corporate credit card invoice.

    Args:
        pdf_total_raw: The raw total value returned by pdf_reader.
        file_name:     Original PDF file name for logging purposes only.

    Returns:
        The validated and rounded PDF total as a float.

    Raises:
        ValueError: If the total is invalid, negative, or unreasonably large.
    """
    if not isinstance(pdf_total_raw, (int, float)):
        raise ValueError(
            f"PDF total must be a number. "
            f"Received: {type(pdf_total_raw)} from '{file_name}'."
        )

    if pdf_total_raw != pdf_total_raw:
        # NaN check: float('nan') != float('nan') is True
        raise ValueError(
            f"PDF total is NaN (not a number) in '{file_name}'. "
            "The PDF may be malformed or the extraction failed."
        )

    if pdf_total_raw < 0:
        raise ValueError(
            f"PDF total is negative (R$ {pdf_total_raw:.2f}) in '{file_name}'. "
            "Invoice totals must be positive values."
        )

    if pdf_total_raw > 1_000_000:
        raise ValueError(
            f"PDF total (R$ {pdf_total_raw:,.2f}) exceeds the safety limit "
            f"of R$ 1,000,000.00. Please verify the file: '{file_name}'."
        )

    if pdf_total_raw == 0.0:
        logger.warning(
            f"PDF total is R$ 0.00 for '{file_name}'. "
            "This may indicate an extraction failure."
        )

    rounded = round(pdf_total_raw, 2)
    logger.debug(f"PDF total validated: R$ {rounded:.2f} | File: {file_name}")
    return rounded


# ─────────────────────────────────────────────
# Function 3 — read_spreadsheet_total
# ─────────────────────────────────────────────

def read_spreadsheet_total(spreadsheet_path: str, sheet_name: str) -> float:
    """
    Opens the spreadsheet and reads the computed SUB TOTAL value
    from the national purchases section of the specified sheet.

    The SUB TOTAL row is identified by scanning column D (DESCRIÇÃO)
    for a cell containing "SUB TOTAL" (case-insensitive). The
    corresponding value is read from column E (VALOR).

    If the cell contains a formula (e.g. =SUM(E5:E54)), openpyxl
    returns None for the cached value when data_only=False. In that
    case, the function manually sums all transaction amounts from
    the data rows as a fallback.

    Args:
        spreadsheet_path: Path to the person's .xlsx file.
        sheet_name:       Name of the worksheet tab.

    Returns:
        The SUB TOTAL value as a float (>= 0.0).

    Raises:
        ValueError: If the file cannot be opened or the sheet is not found.
        ValueError: If no SUB TOTAL row is found in the sheet.
    """
    path = Path(spreadsheet_path)

    if not path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: '{path}'")

    if path.suffix.lower() != ALLOWED_EXTENSION:
        raise ValueError(f"Invalid file type: '{path.suffix}'")

    try:
        # data_only=True reads cached formula results
        workbook = load_workbook(str(path), data_only=True)
    except Exception as exc:
        raise ValueError(
            f"Failed to open spreadsheet '{path.name}': {exc}"
        ) from exc

    if sheet_name not in workbook.sheetnames:
        available = ", ".join(f"'{s}'" for s in workbook.sheetnames)
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: [{available}]"
        )

    worksheet = workbook[sheet_name]
    subtotal_row = _find_subtotal_row(worksheet)

    if subtotal_row is None:
        logger.warning(
            f"SUB TOTAL row not found in '{sheet_name}'. "
            "Falling back to manual sum of transaction rows."
        )
        return _manual_sum_transactions(worksheet)

    raw_value = worksheet.cell(row=subtotal_row, column=COL_AMOUNT).value

    if raw_value is None:
        logger.warning(
            f"SUB TOTAL cell is None (formula not cached) in '{sheet_name}'. "
            "Falling back to manual sum."
        )
        return _manual_sum_transactions(worksheet)

    try:
        total = float(raw_value)
        logger.debug(f"Spreadsheet total read: R$ {total:.2f} | Sheet: {sheet_name}")
        return round(total, 2)
    except (TypeError, ValueError) as exc:
        logger.warning(f"Could not parse SUB TOTAL value '{raw_value}': {exc}")
        return _manual_sum_transactions(worksheet)


# ─────────────────────────────────────────────
# Function 4 — compare_totals
# ─────────────────────────────────────────────

def compare_totals(
    person_name: str,
    pdf_total: float,
    spreadsheet_total: float,
    difference: float,
) -> tuple[str, str]:
    """
    Compares the PDF invoice total against the spreadsheet total
    and determines whether they match within the allowed tolerance.

    A tolerance of R$ 0.01 is applied to handle floating point
    rounding artifacts that can appear when summing BRL values
    (e.g. 1337.61 vs 1337.6099999).

    Args:
        person_name:       Full name of the cardholder.
        pdf_total:         Total extracted from the PDF invoice.
        spreadsheet_total: Total computed from the spreadsheet.
        difference:        Absolute difference between both totals.

    Returns:
        A tuple of (status: str, message: str) where status is one
        of STATUS_MATCH or STATUS_DIVERGENCE, and message is a
        human-readable description of the comparison outcome.
    """
    if not isinstance(pdf_total, (int, float)):
        return STATUS_ERROR, f"Invalid PDF total type: {type(pdf_total)}"

    if not isinstance(spreadsheet_total, (int, float)):
        return STATUS_ERROR, f"Invalid spreadsheet total type: {type(spreadsheet_total)}"

    if not isinstance(difference, (int, float)):
        return STATUS_ERROR, f"Invalid difference type: {type(difference)}"

    is_match = difference <= RECONCILE_TOLERANCE

    if is_match:
        status = STATUS_MATCH
        message = (
            f"✅ MATCH — {person_name}: "
            f"PDF R$ {pdf_total:.2f} = Spreadsheet R$ {spreadsheet_total:.2f} "
            f"(difference: R$ {difference:.2f})"
        )
        logger.info(message)
    else:
        status = STATUS_DIVERGENCE
        message = (
            f"❌ DIVERGENCE — {person_name}: "
            f"PDF R$ {pdf_total:.2f} ≠ Spreadsheet R$ {spreadsheet_total:.2f} "
            f"(difference: R$ {difference:.2f}). Manual review required."
        )
        logger.warning(message)

    return status, message


# ─────────────────────────────────────────────
# Function 5 — report_result
# ─────────────────────────────────────────────

def report_result(result: ReconciliationResult) -> None:
    """
    Prints a formatted reconciliation summary to stdout.

    Output format:
    ─────────────────────────────────────────
    RECONCILIATION RESULT
    ─────────────────────────────────────────
    Person       : Aline Maris
    Sheet        : MAIO 2025
    PDF Total    : R$ 1.337,61
    Sheet Total  : R$ 1.337,61
    Difference   : R$ 0,00
    Status       : ✅ MATCH
    Timestamp    : 2025-05-01T14:30:00
    ─────────────────────────────────────────

    Args:
        result: A fully populated ReconciliationResult dataclass.

    Returns:
        None. Output is printed to stdout only.
    """
    if not isinstance(result, ReconciliationResult):
        logger.error(
            f"report_result received invalid input: {type(result)}. "
            "Expected ReconciliationResult."
        )
        return

    status_icon = "✅" if result.status == STATUS_MATCH else (
                  "❌" if result.status == STATUS_DIVERGENCE else "⚠️")

    separator = "─" * 50

    lines = [
        "",
        separator,
        "  RECONCILIATION RESULT",
        separator,
        f"  Person       : {result.person_name}",
        f"  Sheet        : {result.sheet_name}",
        f"  PDF Total    : R$ {result.pdf_total:,.2f}",
        f"  Sheet Total  : R$ {result.spreadsheet_total:,.2f}",
        f"  Difference   : R$ {result.difference:,.2f}",
        f"  Status       : {status_icon} {result.status}",
        f"  Timestamp    : {result.timestamp}",
        separator,
        "",
    ]

    print("\n".join(lines))
    logger.info(
        f"Result reported | Person: {result.person_name} "
        f"| Status: {result.status} | Diff: R$ {result.difference:.2f}"
    )


# ─────────────────────────────────────────────
# Private helpers
# ─────────────────────────────────────────────

def read_difference(pdf_total: float, spreadsheet_total: float) -> float:
    """
    Calculates the absolute difference between the PDF total
    and the spreadsheet total, rounded to 2 decimal places.

    Args:
        pdf_total:         Total from the PDF invoice.
        spreadsheet_total: Total from the spreadsheet.

    Returns:
        Absolute difference as a float rounded to 2 decimal places.
    """
    if not isinstance(pdf_total, (int, float)) or not isinstance(spreadsheet_total, (int, float)):
        logger.error(
            f"read_difference received invalid types: "
            f"pdf_total={type(pdf_total)}, spreadsheet_total={type(spreadsheet_total)}"
        )
        return 0.0

    difference = abs(round(pdf_total, 2) - round(spreadsheet_total, 2))
    return round(difference, 2)


def write_log(result: ReconciliationResult) -> None:
    """
    Appends a structured JSON log entry to the processing log file.

    Each log entry contains the full ReconciliationResult as JSON,
    one entry per line (JSONL format), making it easy to parse
    programmatically for future reporting or auditing.

    Log directory is created automatically if it does not exist.
    Sensitive fields (passwords, tokens) are never logged.

    Args:
        result: A fully populated ReconciliationResult dataclass.
    """
    DEFAULT_LOG_DIR.mkdir(parents=True, exist_ok=True)

    log_entry = asdict(result)
    log_line = json.dumps(log_entry, ensure_ascii=False)

    try:
        with open(DEFAULT_LOG_FILE, "a", encoding="utf-8") as log_file:
            log_file.write(log_line + "\n")
        logger.debug(f"Log entry written to: {DEFAULT_LOG_FILE}")
    except Exception as exc:
        logger.warning(
            f"Failed to write log entry for '{result.person_name}': {exc}. "
            "Reconciliation result was not persisted to disk."
        )


def _validate_reconcile_inputs(
    person_name: str,
    sheet_name: str,
    spreadsheet_path: str,
    pdf_total: float,
) -> None:
    """
    Validates all inputs to the reconcile() pipeline.

    Raises:
        ValueError: If any required input is missing or invalid.
        FileNotFoundError: If the spreadsheet file does not exist.
    """
    if not person_name or not isinstance(person_name, str) or not person_name.strip():
        raise ValueError(f"person_name must be a non-empty string. Received: {person_name!r}")

    if not sheet_name or not isinstance(sheet_name, str) or not sheet_name.strip():
        raise ValueError(f"sheet_name must be a non-empty string. Received: {sheet_name!r}")

    path = Path(spreadsheet_path)
    if not path.exists():
        raise FileNotFoundError(f"Spreadsheet not found: '{path}'")

    if not isinstance(pdf_total, (int, float)) or pdf_total < 0:
        raise ValueError(
            f"pdf_total must be a non-negative number. Received: {pdf_total!r}"
        )

    logger.debug(
        f"Reconcile input validation passed | person='{person_name}' "
        f"| sheet='{sheet_name}' | pdf_total={pdf_total}"
    )


def _find_subtotal_row(worksheet) -> Optional[int]:
    """
    Scans the worksheet to find the row number of the first
    SUB TOTAL row by checking column D (DESCRIÇÃO) for a
    cell whose value contains "sub total" (case-insensitive).

    Returns:
        The row number (int) if found, or None if not found.
    """
    for row in range(1, MAX_SCAN_ROWS + 1):
        cell_value = worksheet.cell(row=row, column=4).value
        if cell_value is None:
            continue
        cell_str = str(cell_value).strip().lower()
        if any(marker in cell_str for marker in SUBTOTAL_MARKERS):
            logger.debug(f"SUB TOTAL row found at row: {row}")
            return row
    return None


def _manual_sum_transactions(worksheet) -> float:
    """
    Manually sums all numeric values in the VALOR column (E)
    from DATA_START_ROW downward, stopping at the first
    non-numeric, non-empty cell (which marks a summary row).

    This is used as a fallback when the SUB TOTAL formula
    has not been cached by Excel (data_only=True returns None).

    Returns:
        Sum of all transaction amounts as a float.
    """
    DATA_START_ROW = 5
    total = 0.0

    for row in range(DATA_START_ROW, MAX_SCAN_ROWS + 1):
        cell_value = worksheet.cell(row=row, column=COL_AMOUNT).value

        if cell_value is None:
            continue

        try:
            total += float(cell_value)
        except (TypeError, ValueError):
            # Hit a non-numeric cell (formula string or label) — stop summing
            logger.debug(f"Manual sum stopped at row {row} (non-numeric: {cell_value!r})")
            break

    logger.debug(f"Manual sum result: R$ {total:.2f}")
    return round(total, 2)